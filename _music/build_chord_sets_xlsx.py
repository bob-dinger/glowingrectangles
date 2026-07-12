"""Workbook of 2- and 3-chord sets: an INDEX sheet (fill in the NAME per set) +
one tab per set listing every song in the corpus that uses it. Feeds naming.
"""
import os, re, json, psycopg2
from collections import defaultdict
from dotenv import load_dotenv
load_dotenv('/Users/robert/Desktop/themap/themap_claude/.env')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from chord_label import chord_label

def norm(s): return re.sub(r'[^a-z0-9]', '', (s or '').lower())
def core(tok):
    m = re.match(r'^([b#]*[ivxIVX]+(?:/[b#]*[ivxIVX]+)?)', tok)
    return m.group(1) if m else tok
def song_set(hj):
    scale = (hj.get('keys') or [{}])[0].get('scale', 'major')
    scale = scale if scale in ('major', 'minor') else 'major'
    toks = set()
    for c in hj.get('chords') or []:
        r = c.get('root')
        if r and 1 <= r <= 7: toks.add(core(chord_label(c, scale)))
    return frozenset(toks)
def roman_str(s): return ' '.join(sorted(s, key=lambda x: (len(x), x)))

# --- roman token -> chord letter in C major ---
NUM = {'i':1,'ii':2,'iii':3,'iv':4,'v':5,'vi':6,'vii':7}
MAJ_INT = {1:0,2:2,3:4,4:5,5:7,6:9,7:11}
PCNAME = ['C','Db','D','Eb','E','F','Gb','G','Ab','A','Bb','B']
def _plain(tok):
    m = re.match(r'^([b#]*)([ivxIVX]+)$', tok)
    if not m: return None
    acc = sum(-1 if ch == 'b' else 1 for ch in m.group(1))
    num = m.group(2); deg = NUM.get(num.lower())
    if not deg: return None
    pc = (MAJ_INT[deg] + acc) % 12
    qual = 'dim' if (num.lower() == 'vii' and num.islower()) else ('maj' if num.isupper() else 'min')
    return pc, qual
def roman_to_C(tok):
    if '/' in tok:                                    # applied dominant = major
        left, right = tok.split('/', 1)
        rp = _plain(right)
        m = re.match(r'^([b#]*)([ivxIVX]+)$', left)
        if rp and m:
            acc = sum(-1 if ch == 'b' else 1 for ch in m.group(1))
            deg = NUM.get(m.group(2).lower())
            if deg:
                return PCNAME[(rp[0] + MAJ_INT[deg] + acc) % 12]   # major, no suffix
        return tok
    p = _plain(tok)
    if not p: return tok
    pc, q = p
    return PCNAME[pc] + ('m' if q == 'min' else 'dim' if q == 'dim' else '')
def chords_str(s):
    def cpc(c):
        b = c.replace('dim', '').rstrip('m'); return PCNAME.index(b) if b in PCNAME else 99
    return ' '.join(sorted({roman_to_C(t) for t in s}, key=cpc))

# pre-named sets from chord_sets.json (match by roman-token set)
DEG2ROM = {1:'I',2:'ii',3:'iii',4:'IV',5:'V',6:'vi',7:'vii'}
named = {}
for e in json.load(open(os.path.join(os.path.dirname(__file__), 'chord_sets.json')))['entries']:
    if not e.get('ordered') and e.get('degrees'):
        named[frozenset(DEG2ROM[d] for d in e['degrees'] if d in DEG2ROM)] = e['name']

def main():
    c = psycopg2.connect(host=os.environ['DB_HOST'], dbname=os.environ['DB_NAME'], user=os.environ['DB_USER'],
        password=os.environ['DB_PASSWORD'], port=os.environ.get('DB_PORT', 5432))
    cur = c.cursor()
    cur.execute("select artist, title, hookpad_json from parcels.songs where has_chords and hookpad_json is not null")
    rows = cur.fetchall(); c.close()
    seen = set(); by_size = defaultdict(lambda: defaultdict(list))
    for artist, title, hj in rows:
        if not title or title.lower().endswith(('-hooktab', '-simple')): continue
        k = norm((artist or '') + title)
        if k in seen: continue
        seen.add(k)
        s = song_set(hj)
        if 1 <= len(s) <= 8: by_size[len(s)][s].append(f'{artist} - {title}' if artist else title)

    wb = Workbook(); idx = wb.active; idx.title = 'INDEX'
    idx.append(['Size', 'Chords (C major)', 'Set (roman)', '# songs', 'NAME (fill in)', 'tab'])
    for cc in idx[1]:
        cc.font = Font(bold=True, color='FFFFFF'); cc.fill = PatternFill('solid', fgColor='305496')
    used = set()
    def sheetname(size, chords):
        base = f'{size}c {chords}'.replace('/', '_')[:31]
        nm = base; i = 2
        while nm in used: nm = f'{base[:28]}_{i}'; i += 1
        used.add(nm); return nm

    for size in (2, 3):
        for s, songs in sorted(by_size[size].items(), key=lambda kv: -len(kv[1])):
            roman = roman_str(s); chords = chords_str(s); pre = named.get(s, '')
            sn = sheetname(size, chords)
            idx.append([size, chords, roman, len(songs), pre, sn])
            ws = wb.create_sheet(sn)
            ws['A1'] = f'SET:  {chords}   ({roman})'; ws['A1'].font = Font(bold=True, size=13)
            ws['A2'] = 'NAME:'; ws['B2'] = pre
            ws['B2'].fill = PatternFill('solid', fgColor='FFF2CC'); ws['A2'].font = Font(bold=True)
            ws['A3'] = f'{len(songs)} songs'; ws['A3'].font = Font(italic=True, color='888888')
            ws['A5'] = 'Song'; ws['A5'].font = Font(bold=True)
            for i, song in enumerate(sorted(songs), 6): ws[f'A{i}'] = song
            ws.column_dimensions['A'].width = 48
    for col, w in zip('ABCDEF', (6, 20, 20, 9, 26, 16)): idx.column_dimensions[col].width = w
    idx.freeze_panes = 'A2'
    out = os.path.expanduser('~/Desktop/chord_sets_2_3.xlsx')
    wb.save(out)
    print(f'wrote {out}')
    print(f'  {len(by_size[2])} two-chord tabs, {len(by_size[3])} three-chord tabs, +INDEX')

if __name__ == '__main__':
    main()
