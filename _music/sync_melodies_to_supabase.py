"""Push ~/Desktop/melodies_curated.xlsx into parcels.melodies.

Upserts by (slug, section). Rows present in Supabase but not in the xlsx
are listed at the end (not deleted automatically — pass --prune to delete).

Usage:
  python sync_melodies_to_supabase.py
  python sync_melodies_to_supabase.py --prune    # delete orphans
"""
import os, sys
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv('/Users/robert/Desktop/themap/themap_claude/.env')
from supabase import create_client

XLSX = os.path.expanduser('~/Desktop/melodies_curated.xlsx')
sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_SERVICE_ROLE_KEY'])


def main():
    prune = '--prune' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['melodies']
    headers = [c.value for c in ws[1]]
    xlsx_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        d = dict(zip(headers, r))
        xlsx_rows.append({
            'slug':        d['slug'],
            'section':     d['section'],
            'patterns':    d.get('patterns') or '',
            'chord_shape': d.get('chord_shape') or None,
            'notes':       d.get('notes') or None,
        })
    print(f"xlsx: {len(xlsx_rows)} rows")

    # Upsert by (slug, section)
    BATCH = 100
    for i in range(0, len(xlsx_rows), BATCH):
        chunk = xlsx_rows[i:i+BATCH]
        sb.schema('parcels').table('melodies').upsert(
            chunk, on_conflict='slug,section'
        ).execute()
        print(f"  upserted {i+len(chunk)}/{len(xlsx_rows)}")

    # Find orphans
    existing = sb.schema('parcels').table('melodies').select('slug,section').execute().data
    xlsx_keys = {(r['slug'], r['section']) for r in xlsx_rows}
    orphans = [r for r in existing if (r['slug'], r['section']) not in xlsx_keys]
    print(f"\norphans (in Supabase but not in xlsx): {len(orphans)}")
    for o in orphans[:20]:
        print(f"  - {o['slug']} / {o['section']}")
    if orphans and prune:
        for o in orphans:
            sb.schema('parcels').table('melodies').delete().eq('slug', o['slug']).eq('section', o['section']).execute()
        print(f"deleted {len(orphans)} orphans")
    elif orphans:
        print("(pass --prune to delete them)")


if __name__ == '__main__':
    main()
