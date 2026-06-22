"""Build _music/melody-structures.html from ~/Desktop/melodies_curated.xlsx.

For each curated melody, pulls the actual chord + note data from
Supabase so the page can render a per-melody piano roll.

xlsx schema:
  slug, artist, title, section, patterns, chord_shape, notes, hookpad_url
`patterns` is a comma-separated list of pattern strings like
  "8-2222-ABBC", "4-112", "16-4444-AABA"
"""
import os, json, re
from collections import defaultdict
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv('/Users/robert/Desktop/themap/themap_claude/.env')
from supabase import create_client

XLSX = os.path.expanduser('~/Desktop/melodies_curated.xlsx')
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'melody-structures.html')

sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_SERVICE_ROLE_KEY'])


def parse_pattern(p):
    parts = p.strip().split('-')
    if len(parts) < 2: return None
    try: bars = int(parts[0])
    except ValueError: return None
    return (bars, parts[1], parts[2] if len(parts) >= 3 else '')


def norm_name(s):
    return re.sub(r'[^a-z0-9]+', '', (s or '').lower())


def fetch_section_data(slug, section_name):
    """Return {chords, notes, start, end, bpm, key} for the named section, or None."""
    rows = sb.schema('parcels').table('songs').select(
        'slug,hookpad_json'
    ).eq('slug', slug).limit(1).execute().data
    if not rows: return None
    d = rows[0]['hookpad_json'] or {}
    sections = sorted(d.get('sections') or [], key=lambda s: s.get('beat', 0))
    if not sections: return None
    target = norm_name(section_name)
    end_beat = d.get('endBeat') or 0
    bpb = ((d.get('meters') or [{}])[0].get('numBeats')) or 4
    keys = d.get('keys') or [{}]
    tonic = keys[0].get('tonic') or 'C'
    scale = keys[0].get('scale') or 'major'
    tempos = d.get('tempos') or [{}]
    bpm = tempos[0].get('bpm') or 120
    # Find first section whose normalized name matches (exact, then prefix)
    def section_matches(name):
        n = norm_name(name)
        if n == target: return True
        # 'verse' matches 'verse1', 'verse2', etc.
        if n.rstrip('0123456789') == target: return True
        return False
    for i, s in enumerate(sections):
        if section_matches(s.get('name') or s.get('label')):
            start = s.get('beat', 0)
            end = sections[i+1]['beat'] if i+1 < len(sections) else end_beat
            chs = [c for c in (d.get('chords') or [])
                   if c.get('beat') is not None and start <= c['beat'] < end]
            nts = [n for n in (d.get('notes') or [])
                   if n.get('beat') is not None and start <= n['beat'] < end]
            # Compact + zero-based beats
            return {
                'start': start, 'end': end, 'bpb': bpb, 'bpm': bpm,
                'key': tonic, 'scale': scale,
                'chords': [{'r': c.get('root'), 't': c.get('type', 0),
                            'b': round(c['beat'] - start, 3),
                            'd': round(c.get('duration', 1), 3),
                            'app': c.get('applied') or 0,
                            'bor': c.get('borrowed') or ''}
                           for c in chs],
                'notes': [{'sd': n.get('sd'), 'o': n.get('octave', 0),
                           'b': round(n['beat'] - start, 3),
                           'd': round(n.get('duration', 0.5), 3),
                           'r': bool(n.get('isRest'))}
                          for n in nts],
            }
    return None


def build():
    wb = load_workbook(XLSX)
    ws = wb['melodies']
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        rows.append(dict(zip(headers, r)))
    print(f"loaded {len(rows)} curated melodies; fetching section data…")

    # Pull section data for each row (cache by (slug, section))
    cache = {}
    for r in rows:
        key = (r['slug'], r['section'])
        if key not in cache:
            cache[key] = fetch_section_data(r['slug'], r['section'])
        if cache[key] is None:
            print(f"  ! no section data for {r['slug']} / {r['section']}")
        r['_section_data'] = cache[key]

    pat_to_rows = defaultdict(list)
    for r in rows:
        pats_str = (r.get('patterns') or '').strip()
        if not pats_str: continue
        pats = [p.strip() for p in pats_str.split(',') if p.strip()]
        r['_pats_parsed'] = []
        for p in pats:
            parsed = parse_pattern(p)
            if not parsed: continue
            pat_to_rows[p].append(r)
            r['_pats_parsed'].append((p, parsed))

    def pat_sort_key(p):
        parsed = parse_pattern(p)
        if not parsed: return (999, '', '')
        return parsed
    sorted_patterns = sorted(pat_to_rows.keys(), key=pat_sort_key)

    def render_row_data(r):
        return {
            'title': r['title'], 'artist': r['artist'], 'section': r['section'],
            'slug': r['slug'], 'hookpad_url': r.get('hookpad_url') or '',
            'chord_shape': r.get('chord_shape') or '',
            'notes_text': r.get('notes') or '',
            'patterns': [{'str': p, 'bars': parsed[0], 'split': parsed[1], 'letter': parsed[2]}
                         for p, parsed in r.get('_pats_parsed', [])],
            'data': r.get('_section_data'),
        }
    data = {p: [render_row_data(r) for r in pat_to_rows[p]] for p in sorted_patterns}

    sidebar = []
    cur_bars = None
    for p in sorted_patterns:
        bars = parse_pattern(p)[0]
        if bars != cur_bars:
            sidebar.append(f'<div class="bars-header">{bars} bars</div>')
            cur_bars = bars
        sidebar.append(f'<div class="pat-link" data-pat="{p}">{p}<span class="ct">{len(pat_to_rows[p])}</span></div>')

    html = f'''<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Melody Structures — Curated</title>
<style>
  * {{ box-sizing:border-box; }}
  html, body {{ margin:0; height:100%; background:#1a1a2e; color:#e0e0e0; font-family:-apple-system,BlinkMacSystemFont,sans-serif; font-size:13px; overflow:hidden; }}
  .topbar {{ display:flex; align-items:center; gap:14px; padding:10px 16px; border-bottom:1px solid #2a2a4a; background:#0f0f1f; }}
  .topbar a.back {{ color:#6a6a8a; text-decoration:none; font-size:12px; }}
  .topbar a.back:hover {{ color:#e0e0e0; }}
  .topbar h1 {{ font-size:14px; font-weight:700; margin:0; }}
  .topbar .sub {{ color:#6a6a8a; font-size:11px; margin-left:auto; }}
  .layout {{ display:flex; height:calc(100vh - 42px); }}
  .sidebar {{ width:200px; flex-shrink:0; background:#16162a; border-right:1px solid #2a2a4a; overflow-y:auto; padding:6px 0 20px; }}
  .bars-header {{ padding:8px 14px 4px; color:#a5a8fc; font-size:10px; text-transform:uppercase; letter-spacing:1.2px; font-weight:700; background:#1a1a30; margin-top:4px; }}
  .pat-link {{ padding:6px 16px; cursor:pointer; font-family:ui-monospace,Menlo,monospace; font-size:12px; color:#a0a0c0; border-left:3px solid transparent; display:flex; align-items:baseline; }}
  .pat-link:hover {{ background:#22223e; color:#e0e0e0; }}
  .pat-link.active {{ background:#22223e; color:#fff; border-left-color:#3050d0; }}
  .pat-link .ct {{ margin-left:auto; color:#5a5a7a; font-size:10px; }}
  .right {{ flex:1; overflow-y:auto; padding:14px 22px; }}
  .right h2 {{ font-size:14px; color:#e0e0e0; margin:0 0 4px; font-family:ui-monospace,Menlo,monospace; }}
  .right .meta {{ color:#6a6a8a; font-size:11px; margin-bottom:14px; }}
  .pat-blocks {{ display:flex; gap:3px; height:32px; margin-bottom:18px; max-width:540px; }}
  .pat-blocks .ph {{ position:relative; display:flex; align-items:center; justify-content:flex-start; padding-left:8px; border-radius:4px; }}
  .pat-blocks .ph .lt {{ font-weight:700; font-size:12px; color:#fff; text-shadow:0 1px 2px rgba(0,0,0,0.5); }}
  .ph-A {{ background:rgba(60,120,220,0.85); }}
  .ph-B {{ background:rgba(220,90,80,0.85); }}
  .ph-C {{ background:rgba(50,180,80,0.85); }}
  .ph-D {{ background:rgba(230,170,40,0.85); }}
  .ph-E {{ background:rgba(160,80,220,0.85); }}
  .ph-F {{ background:rgba(220,70,200,0.85); }}
  .ph-X {{ background:#3a3a5a; }}
  .card {{ background:#16162a; border:1px solid #2a2a4a; border-radius:6px; padding:10px 14px; margin-bottom:14px; }}
  .card-head {{ display:flex; align-items:baseline; gap:10px; margin-bottom:8px; font-size:13px; flex-wrap:wrap; }}
  .card-head .title {{ font-weight:700; color:#e0e0e0; }}
  .card-head .artist {{ color:#8a8ab0; font-size:12px; }}
  .card-head .section {{ color:#6a6a8a; font-size:11px; }}
  .card-head .key {{ color:#fbbf24; font-size:11px; font-weight:600; }}
  .card-head .hp {{ color:#a5b4fc; text-decoration:none; font-size:11px; font-weight:700; margin-left:auto; padding:2px 8px; background:rgba(99,102,241,0.15); border-radius:4px; }}
  .card-head .hp:hover {{ background:rgba(99,102,241,0.3); }}
  /* piano roll */
  .roll {{ position:relative; height:80px; background:#0d0d1d; border-radius:4px; border:1px solid #2a2a4a; overflow:hidden; }}
  .bar-line {{ position:absolute; top:0; bottom:0; width:1px; background:rgba(140,140,180,0.18); }}
  .bar-line.strong {{ background:rgba(140,140,180,0.45); width:1px; }}
  .chord-bar {{ position:absolute; top:0; height:12px; font-size:9px; color:rgba(255,255,255,0.85); font-family:ui-monospace,Menlo,monospace; padding:0 4px; display:flex; align-items:center; border-radius:2px; }}
  .note {{ position:absolute; height:5px; border-radius:2px; }}
  .n-1 {{ background:#a01e1e; }} .n-2 {{ background:#b35610; }} .n-3 {{ background:#957e0c; }}
  .n-4 {{ background:#25a838; }} .n-5 {{ background:#3050d0; }} .n-6 {{ background:#6e16a5; }}
  .n-7 {{ background:#a01b6b; }}
  .n-s1 {{ background:repeating-linear-gradient(135deg,#a01e1e 0 4px,#b35610 4px 8px); }}
  .n-s2 {{ background:repeating-linear-gradient(135deg,#b35610 0 4px,#957e0c 4px 8px); }}
  .n-s4 {{ background:repeating-linear-gradient(135deg,#25a838 0 4px,#3050d0 4px 8px); }}
  .n-s5 {{ background:repeating-linear-gradient(135deg,#3050d0 0 4px,#6e16a5 4px 8px); }}
  .n-s6 {{ background:repeating-linear-gradient(135deg,#6e16a5 0 4px,#a01b6b 4px 8px); }}
  .n-other {{ background:#888; }}
  .all-patterns {{ font-family:ui-monospace,Menlo,monospace; font-size:10px; color:#8a8ab0; margin-top:8px; }}
  .all-patterns .pat {{ background:#22223e; padding:1px 6px; border-radius:3px; margin-right:5px; cursor:pointer; }}
  .all-patterns .pat:hover {{ background:#3050d0; color:#fff; }}
  .sub-meta {{ font-size:11px; color:#8a8ab0; margin-top:6px; display:flex; gap:14px; flex-wrap:wrap; }}
  .sub-meta .chord {{ color:#fbbf24; font-weight:600; }}
  .sub-meta .notes {{ color:#8a8ab0; font-style:italic; }}
  .placeholder {{ color:#5a5a7a; text-align:center; margin-top:80px; font-size:13px; }}
  .no-data {{ color:#5a5a7a; font-size:11px; font-style:italic; padding:14px; text-align:center; }}
</style>
</head>
<body>
<div class="topbar">
  <a class="back" href="index.html">&larr; Music</a>
  <h1>Melody Structures (curated)</h1>
  <div class="sub">{len(rows)} melodies · {len(sorted_patterns)} unique patterns · source: <code>~/Desktop/melodies_curated.xlsx</code></div>
</div>
<div class="layout">
  <aside class="sidebar">
    {chr(10).join(sidebar)}
  </aside>
  <main class="right" id="right">
    <div class="placeholder">— pick a pattern on the left —</div>
  </main>
</div>

<script>
const DATA = {json.dumps(data, ensure_ascii=False)};

function splitToWidths(split) {{
  if (split.includes('+')) return split.split('+').filter(s => /^\\d+$/.test(s)).map(Number);
  if (/^\\d+$/.test(split)) return split.split('').map(Number);
  return [];
}}

function escapeHtml(s) {{
  return String(s || '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}})[c]);
}}

function renderPatBlocks(pattern) {{
  const widths = splitToWidths(pattern.split);
  const letters = (pattern.letter || '').split('');
  return '<div class="pat-blocks">' + widths.map((w, i) => {{
    const lt = letters[i] || '';
    const lt_norm = lt.toUpperCase().replace(/[^A-F]/g, '').slice(0,1) || 'X';
    return `<div class="ph ph-${{lt_norm}}" style="flex:${{w}}">` +
           (lt ? `<span class="lt">${{escapeHtml(lt)}}</span>` : '') +
           `</div>`;
  }}).join('') + '</div>';
}}

const NUMERALS = ['I','II','III','IV','V','VI','VII'];
function chordLetterShort(c) {{
  const r = c.r;
  if (r < 1 || r > 7) return '?';
  let s = NUMERALS[r-1];
  if (c.app) s = 'V/' + (NUMERALS[c.app-1] || '?');
  if (c.bor) s = 'b' + s;
  if (c.t === 7) s += '7';
  return s;
}}

function sdNum(sd) {{
  if (!sd) return 0;
  if (sd[0] === '#') return parseInt(sd.slice(1)) + 0.5;
  if (sd[0] === 'b') return parseInt(sd.slice(1)) - 0.5;
  return parseInt(sd);
}}
function noteClass(sd) {{
  if (!sd) return 'n-other';
  if (sd[0] === '#') return 'n-s' + sd.slice(1);
  if (sd[0] === 'b') return 'n-s' + (parseInt(sd.slice(1)) - 1);
  return 'n-' + sd;
}}

function renderRoll(d, pattern) {{
  if (!d) return '<div class="no-data">no melody data — slug or section may not match</div>';
  const span = d.end - d.start;
  const bars = Math.round(span / d.bpb);
  // Pitch range from non-rest notes
  const pitches = d.notes.filter(n => !n.r).map(n => n.o * 7 + sdNum(n.sd));
  if (pitches.length === 0) return '<div class="no-data">no notes in this section</div>';
  const lo = Math.min(...pitches), hi = Math.max(...pitches);
  const range = Math.max(1, hi - lo);
  let html = '<div class="roll">';
  // Bar lines
  for (let i = 0; i <= bars; i++) {{
    const left = (i / bars) * 100;
    const cls = (i === 0 || i === bars) ? 'bar-line strong' : 'bar-line';
    html += `<div class="${{cls}}" style="left:${{left}}%"></div>`;
  }}
  // Chord bars across the top
  d.chords.forEach(c => {{
    const left = (c.b / span) * 100;
    const w = (c.d / span) * 100;
    const lbl = chordLetterShort(c);
    let bg = c.r >=1 && c.r <=7 ? `rgba(${{[null,'160,30,30','179,86,16','149,126,12','37,168,56','48,80,208','110,22,165','160,27,107'][c.r]}}, 0.6)` : 'rgba(80,80,100,0.6)';
    html += `<div class="chord-bar" style="left:${{left}}%;width:${{w}}%;background:${{bg}}">${{lbl}}</div>`;
  }});
  // Notes
  d.notes.forEach(n => {{
    if (n.r) return;
    const p = n.o * 7 + sdNum(n.sd);
    // yPct: 0 = top (high), 100 = bottom (low). Use range 18-92 to leave room for chord bars.
    const yPct = 18 + ((hi - p) / range) * 70;
    const left = (n.b / span) * 100;
    const w = Math.max(0.5, (n.d / span) * 100);
    html += `<div class="note ${{noteClass(n.sd)}}" style="left:${{left}}%;width:${{w}}%;top:${{yPct}}%"></div>`;
  }});
  html += '</div>';
  return html;
}}

function show(patStr) {{
  const items = DATA[patStr] || [];
  const right = document.getElementById('right');
  // Find the first item's parsed pattern matching patStr for the header block
  const headerPattern = items.length && items[0].patterns.find(p => p.str === patStr);
  const headerBlocks = headerPattern ? renderPatBlocks(headerPattern) : '';
  const head = `<h2>${{escapeHtml(patStr)}}</h2><div class="meta">${{items.length}} melod${{items.length===1?'y':'ies'}}</div>${{headerBlocks}}`;
  if (!items.length) {{ right.innerHTML = head + '<div class="placeholder">no entries</div>'; return; }}
  const cards = items.map(r => {{
    const thisPat = r.patterns.find(p => p.str === patStr) || r.patterns[0];
    const roll = renderRoll(r.data, thisPat);
    const hp = r.hookpad_url ? `<a class="hp" href="${{r.hookpad_url}}" target="_blank" rel="noopener">HP↗</a>` : '';
    const keyStr = r.data ? `${{r.data.key}} ${{r.data.scale}} · ${{Math.round(r.data.bpm)}}bpm` : '';
    const otherPats = r.patterns.filter(p => p.str !== patStr);
    const allPats = otherPats.length
      ? `<div class="all-patterns">also: ${{otherPats.map(p => `<span class="pat" data-pat="${{escapeHtml(p.str)}}">${{escapeHtml(p.str)}}</span>`).join('')}}</div>`
      : '';
    const extras = [];
    if (r.chord_shape) extras.push(`<span class="chord">${{escapeHtml(r.chord_shape)}}</span>`);
    if (r.notes_text) extras.push(`<span class="notes">${{escapeHtml(r.notes_text)}}</span>`);
    const extraHtml = extras.length ? `<div class="sub-meta">${{extras.join(' · ')}}</div>` : '';
    return `<div class="card">
      <div class="card-head">
        <span class="title">${{escapeHtml(r.title)}}</span>
        <span class="artist">${{escapeHtml(r.artist)}}</span>
        <span class="section">[${{escapeHtml(r.section)}}]</span>
        ${{keyStr ? `<span class="key">${{keyStr}}</span>` : ''}}
        ${{hp}}
      </div>
      ${{roll}}
      ${{allPats}}
      ${{extraHtml}}
    </div>`;
  }}).join('');
  right.innerHTML = head + cards;
  document.querySelectorAll('.pat-link').forEach(p => p.classList.toggle('active', p.dataset.pat === patStr));
  // wire up cross-pattern links inside cards
  right.querySelectorAll('.all-patterns .pat').forEach(el => {{
    el.addEventListener('click', () => show(el.dataset.pat));
  }});
  localStorage.setItem('melodyPat', patStr);
}}

document.querySelectorAll('.pat-link').forEach(p => {{
  p.addEventListener('click', () => show(p.dataset.pat));
}});

const last = localStorage.getItem('melodyPat');
if (last && DATA[last]) show(last);
else {{
  const first = document.querySelector('.pat-link');
  if (first) show(first.dataset.pat);
}}
</script>
</body>
</html>
'''
    with open(OUT, 'w') as f: f.write(html)
    print(f"wrote {OUT}")


if __name__ == '__main__':
    build()
